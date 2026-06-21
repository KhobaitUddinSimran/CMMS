#!/usr/bin/env python3
"""Generate mark import Excel for SMJC4613 Sem 2, 2025/26 — Siti Rahimah"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import random

random.seed(42)  # reproducible

STUDENTS = [
    {"email": "student001@graduate.utm.my",  "full_name": "Ahmad bin Abdullah"},
    {"email": "student002@graduate.utm.my",  "full_name": "Siti Nurhaliza binti Ismail"},
    {"email": "student003@graduate.utm.my",  "full_name": "Muhammad Faisal bin Omar"},
    {"email": "student004@graduate.utm.my",  "full_name": "Nur Aisyah binti Rahman"},
    {"email": "student005@graduate.utm.my",  "full_name": "Lim Wei Jie"},
    {"email": "student006@graduate.utm.my",  "full_name": "Tan Mei Ling"},
    {"email": "student007@graduate.utm.my",  "full_name": "Rajesh Kumar a/l Subramaniam"},
    {"email": "student008@graduate.utm.my",  "full_name": "Priya Sharma"},
    {"email": "student009@graduate.utm.my",  "full_name": "Amirul Hakim bin Saiful"},
    {"email": "student010@graduate.utm.my",  "full_name": "Farah Diba binti Yusuf"},
    {"email": "student011@graduate.utm.my",  "full_name": "Chong Keat Peng"},
    {"email": "student012@graduate.utm.my",  "full_name": "Lee Hui Yin"},
    {"email": "student013@graduate.utm.my",  "full_name": "Vikram Singh"},
    {"email": "student014@graduate.utm.my",  "full_name": "Aishwarya Devi"},
    {"email": "student015@graduate.utm.my",  "full_name": "Hafizuddin bin Mazlan"},
    {"email": "student016@graduate.utm.my",  "full_name": "Nadia binti Kamal"},
    {"email": "student017@graduate.utm.my",  "full_name": "Wong Jun Hao"},
    {"email": "student018@graduate.utm.my",  "full_name": "Kavitha Muthu"},
    {"email": "student019@graduate.utm.my",  "full_name": "Syed Adam bin Syed Ali"},
    {"email": "student020@graduate.utm.my",  "full_name": "Diana binti Rosli"},
]

def generate_excel(output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SMJC4613 Marks"

    # Headers: email | Midterm
    headers = ["email", "Midterm"]
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color="C90031", end_color="C90031", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Generate sample marks (0-100, realistic distribution)
    for s in STUDENTS:
        # Bias toward 50-85 range for realism
        score = round(random.gauss(68, 15))
        score = max(0, min(100, score))  # clamp 0-100
        ws.append([s["email"], score])

    # Adjust column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12

    # Number format for scores
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = "0"

    wb.save(output_path)
    print(f"Mark import Excel saved to: {output_path}")
    print("\nSample scores:")
    for s, score in zip(STUDENTS, [ws.cell(row=i+2, column=2).value for i in range(len(STUDENTS))]):
        print(f"  {s['email']:<40} {score}")

if __name__ == "__main__":
    output = "/Users/khobaituddinsimran/Desktop/ACTIVE WORK/CMMS/SMJC4613_marks_import.xlsx"
    generate_excel(output)
