#!/usr/bin/env python3
"""Create 20 students in the database and generate a roster Excel file."""

import uuid
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os

# Admin password hash — all students will use the same password: password@cmsss
ADMIN_PASSWORD_HASH = "9ab6ebdcf0c4da39a1a31cda11969c57b25b3d76141a3a834b7de964486d2f61"

STUDENTS = [
    {"email": "student001@graduate.utm.my",  "full_name": "Ahmad bin Abdullah",        "matric": "A23MJ0001"},
    {"email": "student002@graduate.utm.my",  "full_name": "Siti Nurhaliza binti Ismail", "matric": "A23MJ0002"},
    {"email": "student003@graduate.utm.my",  "full_name": "Muhammad Faisal bin Omar",    "matric": "A23MJ0003"},
    {"email": "student004@graduate.utm.my",  "full_name": "Nur Aisyah binti Rahman",     "matric": "A23MJ0004"},
    {"email": "student005@graduate.utm.my",  "full_name": "Lim Wei Jie",                 "matric": "A23MJ0005"},
    {"email": "student006@graduate.utm.my",  "full_name": "Tan Mei Ling",                "matric": "A23MJ0006"},
    {"email": "student007@graduate.utm.my",  "full_name": "Rajesh Kumar a/l Subramaniam","matric": "A23MJ0007"},
    {"email": "student008@graduate.utm.my",  "full_name": "Priya Sharma",                "matric": "A23MJ0008"},
    {"email": "student009@graduate.utm.my",  "full_name": "Amirul Hakim bin Saiful",     "matric": "A23MJ0009"},
    {"email": "student010@graduate.utm.my",  "full_name": "Farah Diba binti Yusuf",      "matric": "A23MJ0010"},
    {"email": "student011@graduate.utm.my",  "full_name": "Chong Keat Peng",             "matric": "A23MJ0011"},
    {"email": "student012@graduate.utm.my",  "full_name": "Lee Hui Yin",                 "matric": "A23MJ0012"},
    {"email": "student013@graduate.utm.my",  "full_name": "Vikram Singh",                "matric": "A23MJ0013"},
    {"email": "student014@graduate.utm.my",  "full_name": "Aishwarya Devi",              "matric": "A23MJ0014"},
    {"email": "student015@graduate.utm.my",  "full_name": "Hafizuddin bin Mazlan",       "matric": "A23MJ0015"},
    {"email": "student016@graduate.utm.my",  "full_name": "Nadia binti Kamal",           "matric": "A23MJ0016"},
    {"email": "student017@graduate.utm.my",  "full_name": "Wong Jun Hao",                "matric": "A23MJ0017"},
    {"email": "student018@graduate.utm.my",  "full_name": "Kavitha Muthu",               "matric": "A23MJ0018"},
    {"email": "student019@graduate.utm.my",  "full_name": "Syed Adam bin Syed Ali",      "matric": "A23MJ0019"},
    {"email": "student020@graduate.utm.my",  "full_name": "Diana binti Rosli",           "matric": "A23MJ0020"},
]

def generate_sql():
    now = datetime.utcnow().isoformat()
    sql_lines = []
    sql_lines.append("-- Insert 20 students")
    for s in STUDENTS:
        uid = str(uuid.uuid4())
        sql_lines.append(f"""INSERT INTO users (id, email, full_name, role, password_hash, is_active, email_verified, approval_status, matric_number, special_roles, created_at, updated_at)
VALUES ('{uid}', '{s['email']}', '{s['full_name']}', 'student', '{ADMIN_PASSWORD_HASH}', true, true, 'approved', '{s['matric']}', '{{}}', '{now}', '{now}');""")
    return "\n".join(sql_lines)

def generate_excel(output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Roster"

    # Headers — using UTM format the system accepts
    headers = ["NO.MATRIK", "NAMA", "EMAIL", "SEC"]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="C90031", end_color="C90031", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    sections = ["01", "02", "01", "02", "01", "02", "01", "02", "01", "02",
                "01", "02", "01", "02", "01", "02", "01", "02", "01", "02"]
    for i, s in enumerate(STUDENTS):
        ws.append([s["matric"], s["full_name"], s["email"], sections[i]])

    # Adjust column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 8

    # Save
    wb.save(output_path)
    print(f"Excel roster saved to: {output_path}")

if __name__ == "__main__":
    # Generate SQL
    sql = generate_sql()
    sql_path = "/Users/khobaituddinsimran/Desktop/ACTIVE WORK/CMMS/insert_20_students.sql"
    with open(sql_path, "w") as f:
        f.write(sql)
    print(f"SQL saved to: {sql_path}")

    # Generate Excel
    excel_path = "/Users/khobaituddinsimran/Desktop/ACTIVE WORK/CMMS/student_roster_import.xlsx"
    generate_excel(excel_path)

    print("\nDone! Use the SQL file to insert students into Supabase,")
    print("and use the Excel file for roster import in the CMMS frontend.")
