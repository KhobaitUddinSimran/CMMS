"""Enrollment endpoints - Supabase Edition
Handles student enrollment, roster upload, and add/drop operations.
"""
import logging
import io
import random
import secrets
import string
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from pydantic import BaseModel
from typing import Optional
from core.config import supabase
from core.security import hash_password
from dependencies.auth import get_current_user, has_effective_role
from services.audit_service import AuditService
from services.email_service import EmailService
from utils.session import get_active_session

router = APIRouter(prefix="/api/courses", tags=["enrollments"])
logger = logging.getLogger(__name__)


# ==================== Request Models ====================
class AddStudentRequest(BaseModel):
    student_email: str


# ==================== Helper Functions ====================
def _generate_otp(length: int = 6) -> str:
    """Generate a random numeric OTP (for email verification codes)."""
    return ''.join(random.choices(string.digits, k=length))


def _generate_invitation_token() -> str:
    """Cryptographically strong, URL-safe token used when provisioning student
    accounts from a roster upload. Stored hashed in users.password_hash and
    emailed in plaintext once — the student must use it to set a real
    password on first login."""
    return secrets.token_urlsafe(24)


def _enrollment_session_fields(course: dict) -> dict:
    """Return {semester, academic_year} for enrolling into this course.
    Prefer the course's own session; fall back to the active timeline rather
    than the raw calendar year."""
    sem = course.get("semester")
    ay = course.get("academic_year")
    if sem and ay:
        return {"semester": int(sem), "academic_year": ay}
    active = get_active_session()
    return {
        "semester": int(sem) if sem else active["semester"],
        "academic_year": ay or active["academic_year"],
    }


def _verify_course_access(course: dict, current_user: dict):
    """Verify user has access to the course.
    Admins, coordinators, and HODs have full access.
    Lecturers can only access courses they are assigned to.
    """
    user_role = current_user.get("role")
    user_id = current_user.get("user_id")
    special = set(current_user.get("special_roles", []) or [])
    # Full access for elevated roles (check both role and special_roles for JWT compatibility)
    if user_role in ("admin", "coordinator", "hod") or "coordinator" in special or "hod" in special:
        return
    # Lecturers: scope to their own assigned courses only
    if user_role == "lecturer" and course.get("lecturer_id") != user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You are not assigned to this course"
        )


def _get_course_or_404(course_id: str) -> dict:
    """Fetch course from Supabase or raise 404"""
    response = supabase.table("courses").select("*").eq("id", course_id).execute()
    if not response.data:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Course not found")
    return response.data[0]


# ==================== List Enrolled Students ====================
@router.get("/{course_id}/students")
async def list_enrolled_students(
    course_id: str,
    current_user=Depends(get_current_user),
):
    """Get all students enrolled in a course (joined with user info)"""
    try:
        course = _get_course_or_404(course_id)
        _verify_course_access(course, current_user)

        # Get enrollments for this course
        enroll_resp = (
            supabase.table("enrollments")
            .select("*")
            .eq("course_id", course_id)
            .eq("status", "active")
            .execute()
        )
        enrollments = enroll_resp.data or []

        # Fetch user details for each enrolled student
        students = []
        for e in enrollments:
            user_resp = supabase.table("users").select("id, email, full_name").eq("id", e["student_id"]).execute()
            user = user_resp.data[0] if user_resp.data else {}
            students.append({
                "id": e.get("student_id"),
                "email": user.get("email", ""),
                "full_name": user.get("full_name", ""),
                "enrollment_date": e.get("enrolled_at") or e.get("created_at"),
                "status": e.get("status", "active"),
            })

        return students

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error listing students for course {course_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to list enrolled students")


# ==================== Get Course Enrollments (paginated) ====================
@router.get("/{course_id}/enrollments")
async def get_course_enrollments(
    course_id: str,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    status_filter: Optional[str] = Query(None, alias="status"),
    current_user=Depends(get_current_user),
):
    """List enrollments for a course with pagination"""
    try:
        course = _get_course_or_404(course_id)
        _verify_course_access(course, current_user)

        query = supabase.table("enrollments").select("*").eq("course_id", course_id)
        if status_filter:
            query = query.eq("status", status_filter)

        response = query.range(skip, skip + limit - 1).execute()
        enrollments = response.data or []

        # Get count
        count_query = supabase.table("enrollments").select("id", count="exact").eq("course_id", course_id)
        if status_filter:
            count_query = count_query.eq("status", status_filter)
        count_resp = count_query.execute()
        total = count_resp.count if hasattr(count_resp, "count") and count_resp.count else len(enrollments)

        return {"data": enrollments, "total": total, "skip": skip, "limit": limit}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error listing enrollments for course {course_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to list enrollments")


# ==================== Add Student to Course ====================
@router.post("/{course_id}/enrollments", status_code=status.HTTP_201_CREATED)
async def add_student_to_course(
    course_id: str,
    data: AddStudentRequest,
    current_user=Depends(get_current_user),
):
    """Add a student to a course by email"""
    if not has_effective_role(current_user, "lecturer", "coordinator", "admin"):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        course = _get_course_or_404(course_id)
        _verify_course_access(course, current_user)

        # Find student by email
        user_resp = supabase.table("users").select("*").eq("email", data.student_email).execute()
        if not user_resp.data:
            raise HTTPException(status_code=404, detail="Student not found with that email")

        student = user_resp.data[0]

        # Check not already enrolled
        existing = (
            supabase.table("enrollments")
            .select("id")
            .eq("student_id", student["id"])
            .eq("course_id", course_id)
            .eq("status", "active")
            .execute()
        )
        if existing.data:
            raise HTTPException(status_code=400, detail="Student is already enrolled in this course")

        # Create enrollment with correct session derived from the course /
        # semester_timelines (no more raw datetime.utcnow().year fallback).
        enrollment = {
            "student_id": student["id"],
            "course_id": course_id,
            "status": "active",
            **_enrollment_session_fields(course),
        }
        resp = supabase.table("enrollments").insert(enrollment).execute()

        AuditService.log(
            "ENROLLMENT_CREATED", current_user.get("user_id"), course_id,
            metadata={"student_id": student["id"], "email": student["email"]},
        )

        enrolled_record = resp.data[0] if resp.data else {}
        return {
            "id": student["id"],
            "email": student["email"],
            "full_name": student.get("full_name", ""),
            "enrollment_date": enrolled_record.get("enrolled_at") or enrolled_record.get("created_at"),
            "status": "active",
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error adding student to course {course_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to add student")


# ==================== Drop Student from Course ====================
@router.delete("/{course_id}/enrollments/{student_id}")
async def drop_student_from_course(
    course_id: str,
    student_id: str,
    current_user=Depends(get_current_user),
):
    """Drop (soft-delete) a student from a course"""
    if not has_effective_role(current_user, "lecturer", "coordinator", "admin"):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        course = _get_course_or_404(course_id)
        _verify_course_access(course, current_user)

        # Update enrollment status to dropped
        resp = (
            supabase.table("enrollments")
            .update({"status": "dropped"})
            .eq("student_id", student_id)
            .eq("course_id", course_id)
            .eq("status", "active")
            .execute()
        )

        if not resp.data:
            raise HTTPException(status_code=404, detail="Active enrollment not found")

        return {"message": "Student dropped from course"}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error dropping student {student_id} from course {course_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to drop student")


# ==================== Drop ALL Students from Course ====================
@router.delete("/{course_id}/enrollments")
async def drop_all_students_from_course(
    course_id: str,
    current_user=Depends(get_current_user),
):
    """Drop (soft-delete) ALL active students from a course in one operation"""
    if not has_effective_role(current_user, "lecturer", "coordinator", "admin"):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        course = _get_course_or_404(course_id)
        _verify_course_access(course, current_user)

        resp = (
            supabase.table("enrollments")
            .update({"status": "dropped"})
            .eq("course_id", course_id)
            .eq("status", "active")
            .execute()
        )

        dropped_count = len(resp.data) if resp.data else 0
        return {"message": f"Dropped {dropped_count} student(s) from course", "dropped": dropped_count}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error dropping all students from course {course_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to drop all students")


# ==================== Roster Upload (Preview) ====================
@router.post("/{course_id}/roster/preview")
async def preview_roster_upload(
    course_id: str,
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
):
    """Upload Excel roster file for dry-run preview (no DB changes)"""
    if not has_effective_role(current_user, "lecturer", "coordinator", "hod", "admin"):
        raise HTTPException(status_code=403, detail="Only lecturers, coordinators, and admins can upload rosters")

    try:
        course = _get_course_or_404(course_id)
        _verify_course_access(course, current_user)

        # Validate file
        if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")

        contents = await file.read()
        if len(contents) > 5 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="File size exceeds 5MB limit")

        # Parse Excel
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()

        if len(rows) < 2:
            raise HTTPException(status_code=400, detail="File must have a header row and at least one data row")

        # Parse headers (case-insensitive)
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        
        # Map column indices — supports both legacy format and UTM roster format
        # UTM format: BIL, NO.MATRIK, SEC, NAMA, KUR, PNGK, (empty), T/T
        # Matric is now REQUIRED as primary key. Email is optional.
        col_map = {}
        for i, h in enumerate(headers):
            if h in ("student_id", "matric_number", "matric", "no.matrik", "no. matrik", "no_matrik"):
                col_map["student_id"] = i
            elif h in ("email",):
                col_map["email"] = i
            elif h in ("first_name", "firstname"):
                col_map["first_name"] = i
            elif h in ("last_name", "lastname"):
                col_map["last_name"] = i
            elif h in ("full_name", "name", "nama"):
                col_map["full_name"] = i
            elif h in ("sec", "section"):
                col_map["section"] = i

        # Matric number is REQUIRED - it's the primary lookup key
        if "student_id" not in col_map:
            raise HTTPException(status_code=400, detail="Missing required column: 'NO.MATRIK' or 'matric_number'")

        # Process rows for preview
        students = []
        new_count = 0
        existing_count = 0
        pending_email_count = 0
        error_count = 0

        for row_idx, row in enumerate(rows[1:], start=2):
            try:
                # Skip fully empty rows silently
                if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                # Get matric number (REQUIRED)
                matric = str(row[col_map["student_id"]]).strip() if row[col_map["student_id"]] else ""
                if not matric:
                    continue  # Skip rows without matric

                # Get email (OPTIONAL)
                email = ""
                if "email" in col_map and row[col_map["email"]]:
                    email = str(row[col_map["email"]]).strip().lower()
                    # Validate email domain if provided
                    if email and not (email.endswith("@utm.my") or email.endswith("@graduate.utm.my") or email.endswith("@student.utm.my")):
                        students.append({"student_id": matric, "email": email, "first_name": "", "last_name": "", "status": "error", "error": f"Invalid email domain: {email}"})
                        error_count += 1
                        continue

                # Get name
                if "full_name" in col_map:
                    full_name = str(row[col_map["full_name"]]).strip() if row[col_map["full_name"]] else ""
                    parts = full_name.split(" ", 1)
                    first_name = parts[0]
                    last_name = parts[1] if len(parts) > 1 else ""
                else:
                    first_name = str(row[col_map.get("first_name", 0)]).strip() if col_map.get("first_name") is not None and row[col_map["first_name"]] else ""
                    last_name = str(row[col_map.get("last_name", 0)]).strip() if col_map.get("last_name") is not None and row[col_map["last_name"]] else ""

                section = str(row[col_map["section"]]).strip() if col_map.get("section") is not None and row[col_map["section"]] else ""

                # Check if user exists by matric first (primary key), then email fallback
                existing_user = None
                user_resp = supabase.table("users").select("id, email").eq("matric_number", matric).execute()
                if user_resp.data:
                    existing_user = user_resp.data[0]
                elif email:
                    # Fallback: check by email for backwards compatibility
                    user_resp = supabase.table("users").select("id, email").eq("email", email).execute()
                    if user_resp.data:
                        existing_user = user_resp.data[0]

                if existing_user:
                    students.append({"student_id": matric, "email": email or existing_user.get("email", ""), "first_name": first_name, "last_name": last_name, "section": section, "status": "existing"})
                    existing_count += 1
                else:
                    # New student - check if email provided
                    if email:
                        students.append({"student_id": matric, "email": email, "first_name": first_name, "last_name": last_name, "section": section, "status": "new"})
                        new_count += 1
                    else:
                        students.append({"student_id": matric, "email": "", "first_name": first_name, "last_name": last_name, "section": section, "status": "pending_email"})
                        pending_email_count += 1

            except Exception as row_err:
                students.append({"student_id": "", "email": "", "first_name": "", "last_name": "", "status": "error", "error": f"Row {row_idx}: {str(row_err)}"})
                error_count += 1

        return {
            "students": students,
            "new_count": new_count,
            "existing_count": existing_count,
            "pending_email_count": pending_email_count,
            "error_count": error_count,
            "summary": f"{new_count} new accounts will be created, {existing_count} existing linked, {pending_email_count} pending email",
            "total_rows": len(students),
        }

    except HTTPException:
        raise
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")
    except Exception as e:
        logger.error(f"Error previewing roster for course {course_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to preview roster")


# ==================== Roster Upload (Actual Import) ====================
@router.post("/{course_id}/roster/upload")
async def upload_roster(
    course_id: str,
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
):
    """Upload Excel roster and create accounts + enrollments.
    
    For each student row:
    - If email exists → link enrollment
    - If email is new → create user with OTP, send email, link enrollment
    """
    if not has_effective_role(current_user, "lecturer", "coordinator", "hod", "admin"):
        raise HTTPException(status_code=403, detail="Only lecturers, coordinators, and admins can upload rosters")

    try:
        course = _get_course_or_404(course_id)
        _verify_course_access(course, current_user)

        # Validate and parse file
        if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")

        contents = await file.read()
        if len(contents) > 5 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="File size exceeds 5MB limit")

        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()

        if len(rows) < 2:
            raise HTTPException(status_code=400, detail="File must have header + data rows")

        # Parse headers — supports both legacy and UTM roster format
        # UTM format: BIL, NO.MATRIK, SEC, NAMA, KUR, PNGK, (empty), T/T
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        col_map = {}
        for i, h in enumerate(headers):
            if h in ("student_id", "matric_number", "matric", "no.matrik", "no. matrik", "no_matrik"):
                col_map["student_id"] = i
            elif h in ("email",):
                col_map["email"] = i
            elif h in ("first_name", "firstname"):
                col_map["first_name"] = i
            elif h in ("last_name", "lastname"):
                col_map["last_name"] = i
            elif h in ("full_name", "name", "nama"):
                col_map["full_name"] = i
            elif h in ("sec", "section"):
                col_map["section"] = i

        # Matric number is REQUIRED as primary lookup key. Email is optional.
        if "student_id" not in col_map:
            raise HTTPException(status_code=400, detail="Missing required column: 'NO.MATRIK' or 'matric_number'")

        created_count = 0
        linked_count = 0
        pending_email_count = 0
        errors = []
        student_emails_created = []

        for row_idx, row in enumerate(rows[1:], start=2):
            try:
                # Skip fully empty rows silently
                if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                # Get matric number (REQUIRED)
                matric = str(row[col_map["student_id"]]).strip() if row[col_map["student_id"]] else ""
                if not matric:
                    continue

                # Get email (OPTIONAL)
                email = ""
                if "email" in col_map and row[col_map["email"]]:
                    email = str(row[col_map["email"]]).strip().lower()
                    # Validate email domain if provided
                    if email and not (email.endswith("@utm.my") or email.endswith("@graduate.utm.my") or email.endswith("@student.utm.my")):
                        errors.append(f"Row {row_idx}: Invalid email domain: {email}")
                        continue

                # Get name
                if "full_name" in col_map:
                    full_name = str(row[col_map["full_name"]]).strip() if row[col_map["full_name"]] else matric
                    parts = full_name.split(" ", 1)
                    first_name = parts[0]
                    last_name = parts[1] if len(parts) > 1 else ""
                else:
                    first = str(row[col_map.get("first_name", 0)]).strip() if col_map.get("first_name") is not None and row[col_map["first_name"]] else ""
                    last = str(row[col_map.get("last_name", 0)]).strip() if col_map.get("last_name") is not None and row[col_map["last_name"]] else ""
                    full_name = f"{first} {last}".strip() or matric
                    first_name = first
                    last_name = last

                # Check if user exists by matric first (primary key), then email fallback
                existing_user = None
                user_resp = supabase.table("users").select("id, email, matric_number").eq("matric_number", matric).execute()
                if user_resp.data:
                    existing_user = user_resp.data[0]
                elif email:
                    # Fallback: check by email for backwards compatibility
                    user_resp = supabase.table("users").select("id, email, matric_number").eq("email", email).execute()
                    if user_resp.data:
                        existing_user = user_resp.data[0]
                        # If found by email but no matric, update the matric
                        if not existing_user.get("matric_number"):
                            supabase.table("users").update({"matric_number": matric}).eq("id", existing_user["id"]).execute()

                if existing_user:
                    # Existing user → just enroll
                    student_id = existing_user["id"]
                    linked_count += 1
                else:
                    # New user → provision as an invited, not-yet-active account.
                    # Generate a one-time invitation token that the student will
                    # exchange for a real password on first login.
                    invite_token = _generate_invitation_token()
                    hashed = hash_password(invite_token)
                    expires_at = (datetime.utcnow() + timedelta(days=14)).isoformat()

                    new_user = {
                        "matric_number": matric,
                        "full_name": full_name,
                        "password_hash": hashed,
                        "role": "student",
                        "is_active": False,  # activates when invite is accepted
                        "email_verified": False,
                        "approval_status": "approved",  # coordinator-provisioned
                        "invitation_status": "pending" if email else "pending_email",  # no email = needs manual invite
                        "invitation_expires_at": expires_at if email else None,
                        "invited_by": current_user.get("user_id"),
                    }
                    # Only set email if provided
                    if email:
                        new_user["email"] = email

                    create_resp = supabase.table("users").insert(new_user).execute()
                    if not create_resp.data:
                        errors.append(f"Row {row_idx}: Failed to create account for matric {matric}")
                        continue

                    student_id = create_resp.data[0]["id"]
                    created_count += 1
                    if email:
                        student_emails_created.append(email)
                    else:
                        pending_email_count += 1

                    # Send invitation email only if email was provided
                    if email:
                        try:
                            await EmailService.send_student_otp(email, invite_token, full_name)
                        except Exception as email_err:
                            logger.error(f"Failed to send invitation email to {email}: {email_err}")
                            errors.append(f"Row {row_idx}: Account created but invitation email failed for {email} — resend from Roster Management")

                # Create enrollment (skip if already active) with correct session
                existing_enroll = (
                    supabase.table("enrollments")
                    .select("id")
                    .eq("student_id", student_id)
                    .eq("course_id", course_id)
                    .eq("status", "active")
                    .execute()
                )
                if not existing_enroll.data:
                    supabase.table("enrollments").insert({
                        "student_id": student_id,
                        "course_id": course_id,
                        "status": "active",
                        **_enrollment_session_fields(course),
                    }).execute()

            except Exception as row_err:
                errors.append(f"Row {row_idx}: {str(row_err)}")

        AuditService.log(
            "ROSTER_UPLOADED", current_user.get("user_id"), course_id,
            metadata={
                "created": created_count,
                "linked": linked_count,
                "errors": len(errors),
                "total": created_count + linked_count,
            },
        )

        return {
            "success": len(errors) == 0,
            "accounts_created": created_count,
            "enrollments_linked": linked_count,
            "pending_email_count": pending_email_count,
            "errors": len(errors),
            "error_details": errors,
            "message": f"Roster processed: {created_count} accounts created, {linked_count} existing linked, {pending_email_count} pending email, {len(errors)} errors",
            "student_emails": student_emails_created,
            "total_rows": created_count + linked_count + pending_email_count + len(errors),
            "created_students": created_count,
            "existing_students": linked_count,
            "pending_students": pending_email_count,
            "failed_students": len(errors),
        }

    except HTTPException:
        raise
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")
    except Exception as e:
        logger.error(f"Error uploading roster for course {course_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to upload roster")


# ==================== Roster Confirm (alias for upload) ====================
@router.post("/{course_id}/roster/confirm")
async def confirm_roster_import(
    course_id: str,
    current_user=Depends(get_current_user),
):
    """Confirm a previously previewed roster import. 
    In this implementation, the actual upload endpoint handles everything,
    so this just returns a success acknowledgement."""
    return {"message": "Import confirmed", "status": "complete"}


# ==================== Invite Pending Student ====================
class InvitePendingStudentRequest(BaseModel):
    email: str


@router.post("/{course_id}/students/{student_id}/invite")
async def invite_pending_student(
    course_id: str,
    student_id: str,
    data: InvitePendingStudentRequest,
    current_user=Depends(get_current_user),
):
    """Add email to a pending student (created without email) and send invitation.
    
    This allows coordinators to manually invite students who were imported
    from rosters that didn't include email addresses.
    """
    if not has_effective_role(current_user, "lecturer", "coordinator", "hod", "admin"):
        raise HTTPException(status_code=403, detail="Only lecturers, coordinators, and admins can invite students")
    
    try:
        # Validate course access
        course = _get_course_or_404(course_id)
        _verify_course_access(course, current_user)
        
        # Validate email format
        email = data.email.strip().lower()
        if not email or "@" not in email:
            raise HTTPException(status_code=400, detail="Invalid email format")
        
        # Check if email already exists
        existing = supabase.table("users").select("id").eq("email", email).execute()
        if existing.data:
            raise HTTPException(status_code=400, detail="Email already registered to another student")
        
        # Get the pending student
        student_resp = supabase.table("users").select("*").eq("id", student_id).eq("role", "student").execute()
        if not student_resp.data:
            raise HTTPException(status_code=404, detail="Student not found")
        
        student = student_resp.data[0]
        
        # Verify student is in pending state (no email or invitation_status is pending_email)
        if student.get("email") and student.get("invitation_status") != "pending_email":
            raise HTTPException(status_code=400, detail="Student already has an email and invitation")
        
        # Generate invitation token
        invite_token = _generate_invitation_token()
        hashed = hash_password(invite_token)
        expires_at = (datetime.utcnow() + timedelta(days=14)).isoformat()
        
        # Update student with email and invitation
        update_data = {
            "email": email,
            "password_hash": hashed,
            "invitation_status": "pending",
            "invitation_expires_at": expires_at,
            "invited_by": current_user.get("user_id"),
        }
        
        update_resp = supabase.table("users").update(update_data).eq("id", student_id).execute()
        if not update_resp.data:
            raise HTTPException(status_code=500, detail="Failed to update student email")
        
        # Send invitation email
        try:
            await EmailService.send_student_otp(email, invite_token, student.get("full_name", ""))
        except Exception as email_err:
            logger.error(f"Failed to send invitation email to {email}: {email_err}")
            raise HTTPException(status_code=500, detail=f"Email updated but invitation failed to send: {str(email_err)}")
        
        AuditService.log(
            "STUDENT_INVITED", current_user.get("user_id"), student_id,
            metadata={"course_id": course_id, "email": email, "matric": student.get("matric_number")},
        )
        
        return {
            "success": True,
            "message": f"Invitation sent to {email}",
            "student_id": student_id,
            "email": email,
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error inviting student {student_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to send invitation")
